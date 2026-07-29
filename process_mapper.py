#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import re
import sqlite3
from dataclasses import dataclass
from typing import Optional

SYNONYM_MAP = {
    # 영문 약어
    "RC": "콘크리트", "레미콘": "콘크리트", "무근": "무근콘크리트",
    "철근콘크리트": "콘크리트", "PC": "프리캐스트",
    "철골": "강구조", "스틸": "강구조", "H빔": "강구조",
    "데크": "데크플레이트", "내화뿜칠": "내화피복",
    # NOTE: "방수", "타일", "도장" 등 이미 공종명의 핵심어인 경우
    # "공사"를 붙이면 "방수공사공사" 가 되므로 제외.
    # 대신 threshold를 낮춰 직접 매칭.
    "ALC블록": "ALC블록", "ALC": "ALC블록",
    "도막방수": "도막방수", "시트방수": "시트방수",
    "우레탄방수": "도막방수", "벤토나이트": "벤토나이트방수",
    "실링": "실링", "코킹": "실링",
    "외단열": "외단열", "EIFS": "외단열",
    "결로방지": "결로방지",
    "모르타르": "시멘트모르타르", "에폭시바닥": "합성고분자바닥",
    "화강석": "화강석", "대리석": "대리석", "건식석재": "건식석재",
    "커튼월": "커튼월",
    "알루미늄창호": "알루미늄창호",
    "PVC창호": "합성수지창호",
    "스테인리스창호": "스테인리스창호",
    "싱글": "아스팔트싱글", "금속지붕": "금속판지붕",
    "도배": "도배", "마루": "마루", "온돌": "온돌",
    "난간": "금속난간",
    "내화충전": "내화충전", "방화구획": "내화충전",
    "ALC패널": "ALC패널",
    "철거": "해체",
    "버티컬": "블라인드",
    "페인트": "도장",
    # 토공사 계열 → 건축물 부대공사 키워드로 유도
    "토공사": "배수잡시설", "토및지정": "배수잡시설",
    "부지정리": "배수잡시설", "굴착": "배수잡시설",
    "정화조": "정화조", "오수": "오수정화",
    "배수": "배수", "굴뚝": "굴뚝",
    "담장": "담장울타리", "대문": "담장울타리", "울타리": "담장울타리",
}

SPEC_ONLY_KEYWORDS = [
    "프로젝터", "빔프로젝터", "스크린", "음향", "조명기구",
    "가구", "붙박이장", "주방기구", "싱크대", "욕실기구",
    "거울설치", "액자", "사인", "간판",
    "엘리베이터", "에스컬레이터", "리프트",
    "소화기", "소화설비", "스프링클러",
    "CCTV", "자동문",
    "에어컨", "냉난방기", "FCU",
    "태양광", "ESS",
    "주차설비", "턴테이블",
]


@dataclass
class MappingResult:
    input_name: str
    status: str
    detail_code: Optional[str] = None
    detail_name: Optional[str] = None
    process_code: Optional[str] = None
    process_name: Optional[str] = None
    score: float = 0.0
    note: str = ""


class ProcessMapper:
    def __init__(self, db_path: str):
        self.db_path = db_path
        self._load_index()

    def _load_index(self):
        con = sqlite3.connect(self.db_path)
        rows = con.execute(
            "SELECT process_code, process_name, detail_code, detail_name FROM specs"
        ).fetchall()
        con.close()
        self.index = [
            {
                "process_code": r[0],
                "process_name": r[1] or "",
                "detail_code": r[2],
                "detail_name": r[3],
                # detail_name을 두 번 넣어 가중치 높임 (process_name은 보조)
                "tokens": self._tokenize(f"{r[3]} {r[3]} {r[1] or ''}"),
            }
            for r in rows
        ]

    @staticmethod
    def _tokenize(text: str) -> set:
        text = re.sub(r"[^\w가-힣]", " ", text).lower()
        tokens = set(t for t in text.split() if len(t) > 1)
        joined = re.sub(r"\s+", "", text)
        for n in (2, 3, 4):
            for i in range(len(joined) - n + 1):
                tokens.add(joined[i:i+n])
        return tokens

    @staticmethod
    def _normalize(text: str) -> str:
        for alias, canonical in SYNONYM_MAP.items():
            text = re.sub(re.escape(alias), canonical, text, flags=re.IGNORECASE)
        return text

    def _score(self, query_tokens: set, item: dict) -> float:
        if not query_tokens or not item["tokens"]:
            return 0.0
        intersection = query_tokens & item["tokens"]
        weighted = sum(len(t) for t in intersection)
        total = sum(len(t) for t in query_tokens | item["tokens"])
        return weighted / total if total else 0.0

    def map(self, input_name: str, threshold_standard: float = 0.15,
            threshold_ref: float = 0.05) -> MappingResult:
        # 앞에 붙은 번호 제거: "01.", "1.", "1-" 형식
        cleaned = re.sub(r'^\d+[\.\-]\s*', '', input_name).strip()
        # 한글 낱자 공백 정규화: "가 설 공 사" → "가설공사"
        cleaned = _normalize_korean(cleaned)

        lower = cleaned.lower()
        for kw in SPEC_ONLY_KEYWORDS:
            if kw.lower() in lower:
                return MappingResult(
                    input_name=input_name,
                    status="SPEC_ONLY",
                    note="KCS 표준시방서 해당 없음 -> 특기시방서(AI초안) 생성 대상",
                )

        normalized = self._normalize(cleaned)
        query_tokens = self._tokenize(normalized)

        scored = sorted(
            [(self._score(query_tokens, item), idx) for idx, item in enumerate(self.index)],
            reverse=True
        )
        if not scored:
            return MappingResult(input_name=input_name, status="SPEC_ONLY",
                                 note="인덱스 없음")
        best_score, best_idx = scored[0]
        best = self.index[best_idx]

        if best_score < threshold_ref:
            return MappingResult(input_name=input_name, status="SPEC_ONLY",
                                 note="KCS 유사 항목 없음")
        elif best_score >= threshold_standard:
            return MappingResult(
                input_name=input_name, status="STANDARD",
                detail_code=best["detail_code"], detail_name=best["detail_name"],
                process_code=best["process_code"], process_name=best["process_name"],
                score=round(best_score, 3),
            )
        else:
            return MappingResult(
                input_name=input_name, status="STANDARD_REF",
                detail_code=best["detail_code"], detail_name=best["detail_name"],
                process_code=best["process_code"], process_name=best["process_name"],
                score=round(best_score, 3),
                note=f"근접 참조(유사도 {round(best_score,3)}) -> 특기시방서에서 구체화 필요",
            )

    def map_list(self, names):
        return [self.map(n) for n in names]

    def all_specs(self):
        return [
            {"process_code": item["process_code"], "process_name": item["process_name"],
             "detail_code": item["detail_code"], "detail_name": item["detail_name"]}
            for item in self.index
        ]

    def search(self, query: str, top_k: int = 10):
        normalized = self._normalize(query)
        query_tokens = self._tokenize(normalized)
        scored = sorted(
            [(self._score(query_tokens, item), idx) for idx, item in enumerate(self.index)],
            reverse=True
        )
        results = []
        for score, idx in scored[:top_k]:
            if score > 0:
                item = self.index[idx]
                results.append({**item, "score": round(score, 3)})
        return results


def _normalize_korean(text: str) -> str:
    """한글 낱자 사이 공백 제거: '가 설 공 사' → '가설공사'"""
    # 한글 한 글자씩 공백으로 분리된 패턴 정규화
    text = re.sub(r'(?<=[가-힣])\s(?=[가-힣])', '', text)
    return text.strip()


def _extract_from_sheet(ws) -> list:
    """
    시트에서 공종명 목록 반환.
    - 6자리 코드(010101~010119)가 있는 행을 찾아서
    - 같은 행의 한글 셀 값을 공종명으로 사용.
    - 코드+이름이 같은 셀에 있거나 별도 셀에 있는 경우 모두 처리.
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    names = []

    for row in rows:
        # 모든 셀을 문자열로 변환
        cells = [str(c).strip() if c is not None else '' for c in row]

        # 이 행에 6자리 숫자 코드가 있는지 확인
        found_6digit = False
        for cell in cells:
            # "010102", "010102 가설공사", 또는 숫자 10102 형태 모두 처리
            digits = re.sub(r'\D', '', cell)
            if len(digits) == 6:
                found_6digit = True
                break
            # 선행 0이 빠진 5자리 숫자 (Excel이 10102로 저장한 경우)
            if len(digits) == 5 and cell.startswith(digits):
                found_6digit = True
                break

        if not found_6digit:
            continue

        # 같은 행에서 한글이 포함된 셀을 공종명으로 사용
        for cell in cells:
            if re.search(r'[가-힣]', cell):
                # 코드가 앞에 붙어 있으면 분리: "010102 가설공사" → "가설공사"
                name = re.sub(r'^\d+\s*', '', cell).strip()
                if not name:
                    name = cell.strip()
                normalized = _normalize_korean(name)
                if len(normalized) > 1:
                    names.append(normalized)
                break

    return list(dict.fromkeys(n for n in names if len(n) > 1))


def extract_from_excel(file_path: str, detail: bool = False) -> list:
    """
    엑셀 내역서에서 공종명 추출.

    Parameters
    ----------
    file_path : str
        엑셀 파일 경로
    detail : bool
        True이면 '공종별내역서' 시트도 함께 참조 (세부 공종 추가)

    Returns
    -------
    list[str]  공종명 목록
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet_names = wb.sheetnames

        # 우선순위: 공종별집계표 > 공종별내역서 > active sheet
        SUMMARY_KEYWORDS = ["공종별집계표", "집계표"]
        DETAIL_KEYWORDS  = ["공종별내역서", "내역서"]

        def find_sheet(keywords):
            for kw in keywords:
                for sn in sheet_names:
                    if kw in sn:
                        return wb[sn]
            return None

        summary_ws = find_sheet(SUMMARY_KEYWORDS)
        detail_ws  = find_sheet(DETAIL_KEYWORDS) if detail else None

        if summary_ws:
            names = _extract_from_sheet(summary_ws)
            if detail and detail_ws:
                detail_names = _extract_from_sheet(detail_ws)
                # 세부 공종 추가 (중복 제외)
                existing = set(names)
                for n in detail_names:
                    if n not in existing:
                        names.append(n)
                        existing.add(n)
            return names

        # 집계표 시트를 못 찾은 경우 active sheet 사용
        return _extract_from_sheet(wb.active)

    except Exception:
        return []


def get_excel_sheets(file_path: str) -> list:
    """엑셀 파일의 시트 목록 반환"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True)
        return wb.sheetnames
    except Exception:
        return []


def extract_from_text(text: str):
    text = re.sub(r"\d+\.\s*", "", text)
    items = re.split(r"[,\n;]+", text)
    return [i.strip() for i in items if i.strip() and len(i.strip()) > 1]


if __name__ == "__main__":
    import sys, os
    db_path = sys.argv[1] if len(sys.argv) > 1 else "output/specs.db"
    if not os.path.exists(db_path):
        print(f"DB 없음: {db_path}")
        sys.exit(1)
    mapper = ProcessMapper(db_path)
    tests = ["철골공사", "콘크리트", "외단열", "도장", "커튼월",
             "타일", "방수공사", "창호", "거울설치", "프로젝터설치",
             "블라인드", "에폭시바닥", "내화충전", "해체공사"]
    print(f"{'입력':<20} {'상태':<14} {'코드':<16} {'매핑명'}")
    print("-" * 80)
    for name in tests:
        r = mapper.map(name)
        code = r.detail_code or "-"
        mname = r.detail_name if r.detail_name else r.note[:30]
        print(f"{name:<20} {r.status:<14} {code:<16} {mname}")
